#!/usr/bin/env python3
"""
S&P 500 Dashboard Updater v5.0 — Iran War / Oil Shock Refresh (May 22, 2026)
Fetches live financial data from Yahoo Finance and updates the Dashboard sheet in Portfolio.xlsx.
Run manually via the "Update Data" button in the Dashboard sheet, or from terminal.

Features:
- Batch fetching with automatic retries and exponential backoff
- Per-ticker error isolation (one failure doesn't block others)
- Detailed error log written to 'Errors' sheet
- Data validation: sanity checks on prices, ratios, market caps
- Buffett-style intrinsic value: 15-yr DCF with decaying growth + liquid assets + margin of safety
- GEOPOLITICAL RISK OVERLAY v5 (May 22, 2026): IRAN WAR + OIL SHOCK INTENSIFIES
  * Iran war ONGOING since Feb 28 (Op Epic Fury killed Khamenei). Brent SPIKED
    to $107-112 (Apr 8 ceasefire holding but fragile, oil +45% since Feb).
    WTI $102-108. Hormuz mostly closed by Iran, US blockading Iranian ports.
    May 4 Project Freedom escort launched; May 6 paused; May 18 Trump called
    off planned Tuesday strike; Pakistan mediating Iran's 14-point framework.
  * Supreme Court ruled IEEPA tariff authority unconstitutional (Feb 2026).
    Tariffs now ~10% (down from 11%), but other statutes preserve broad
    presidential authority. ISM prices index 84.6 (highest since Apr 2022)
    on tariff + energy double pressure.
  * Fed 3.50-3.75% (held Apr 28-29, 8-4 vote with 4 dissents). PCE forecast
    2.7% — sticky. Job creation near zero past year (unusual outside recession).
    Markets price 1-2 cuts H2-26 but inflation fight constrains Fed.
  * Recession risk: Goldman baseline still no recession (2.3-2.6% GDP), but
    STAGFLATION re-emerging on oil shock + tariff persistence. ~28% WSJ.
  * S&P 500: New ATH 7,501 (May 15 close, +0.8%). Forward P/E 20.9-21.2x.
    Q1 blended earnings +15.1% (beat). 2026 EPS growth fcst 18.6%.
  * BRK Q1 2026 13F (filed May 15): $263B portfolio, 29 positions (-13).
    Cash $380-397B. 16 exits incl V/MA/UNH/DPZ/CHTR/HEI/LAMR/POOL/AON/ALLE/DEO.
    CVX -35% (-46M sh, sold ~$8.4B). STZ -95%. GOOGL TRIPLED to $15.6B.
    New positions: DAL ($2.6B), GOOG ($1B), Macy's ($55M). NYT tripled.
    Cash war chest $397B (highest ever). $234M buybacks (first since May-24).
  * Sector-specific: Energy/Defense/Insurance STRONG beneficiaries (Brent $112);
    Transportation/Auto/Industrial HEAVILY hurt (fuel + tariffs); Tech mild
    valuation premium; consumer pain RE-INTENSIFIED by oil + tariffs.
- JSON summary report saved alongside for diagnostics

Required: pip install yfinance openpyxl pandas
"""

import sys
import os
import json
import time
import datetime
import traceback
import warnings
warnings.filterwarnings('ignore')

# ── Dependency check ──────────────────────────────────────────────────────────
REQUIRED = ['yfinance', 'pandas', 'openpyxl']
missing = []
for pkg in REQUIRED:
    try:
        __import__(pkg)
    except ImportError:
        missing.append(pkg)
if missing:
    print(f"Missing packages: {', '.join(missing)}")
    print(f"Installing...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', *missing, '-q'])
    print("Installed. Re-importing...")

import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, 'Portfolio.xlsx')
LOG_PATH = os.path.join(SCRIPT_DIR, 'dashboard_update.log')
REPORT_PATH = os.path.join(SCRIPT_DIR, 'last_update_report.json')
BATCH_SIZE = 40
MAX_RETRIES = 3
BASE_DELAY = 3  # seconds, doubles each retry

HEADER_ROW = 5
DATA_START = 6

# ── Validation bounds ─────────────────────────────────────────────────────────
VALID_RANGES = {
    'price':        (0.01, 100_000),
    'mcap_b':       (0.001, 20_000),
    'pe':           (-5000, 5000),
    'gross_margin': (-5, 5),
    'rev_growth':   (-10, 100),
    'fcf_m':        (-500_000, 500_000),
    'p_fcf':        (-5000, 5000),
}

def validate(key, val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    lo, hi = VALID_RANGES.get(key, (None, None))
    if lo is not None and not (lo <= val <= hi):
        return None
    return val


# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg):
    ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    try:
        with open(LOG_PATH, 'a') as f:
            f.write(line + '\n')
    except Exception:
        pass



# ── Geopolitical risk overlay v5 (May 22, 2026) — IRAN WAR + OIL SHOCK ──────
# Current macro environment:
#   - Iran war ONGOING since Feb 28, 2026 (US/Israel Operation Epic Fury killed
#     Supreme Leader Khamenei). Apr 8 ceasefire HOLDING but FRAGILE.
#     Brent SPIKED to $107-112 (May 18-21); WTI $102-108. Oil up ~45% since
#     Iran war began. Hormuz MOSTLY CLOSED by Iran; US blockading Iranian ports.
#     May 4 Project Freedom escort op launched → paused May 6 → May 18 Trump
#     called off planned Tuesday strike. Pakistan mediating Iran's 14-point
#     framework. Iran reviewing US position May 21.
#     Chubb (CB) lead underwriter for US Hormuz shipping insurance program.
#   - Supreme Court ruled IEEPA tariff authority UNCONSTITUTIONAL (Feb 2026):
#     overall tariffs ~10% (down from 11%), but other statutes preserve broad
#     presidential authority — tariffs unlikely to disappear. ISM prices index
#     84.6 (highest since Apr 2022) on tariff + energy DOUBLE pressure.
#   - Fed: 3.50-3.75% held Apr 28-29 — 8-4 vote, 4 DISSENTS. PCE forecast 2.7%
#     — sticky. Job creation near zero past year (unusual outside recession).
#     Markets price 1-2 cuts H2-26 but inflation fight constrains Fed.
#   - Recession risk: Goldman baseline still no recession (2.3-2.6% GDP), but
#     STAGFLATION re-emerging on oil shock + tariff persistence. ~28% WSJ.
#   - S&P 500: NEW ATH 7,501 close May 15 (+0.8%, intraday 7,517). Forward
#     P/E ~20.9-21.2x. Q1 blended earnings +15.1% (beat). 2026 EPS growth
#     forecast 18.6%. Nasdaq new record 26,635. AI semis driving.
#   - BRK Q1 2026 13F (filed May 15): $263B portfolio, 29 positions (-13).
#     Cash $380-397B (record). 16 exits incl V/MA/UNH/DPZ/CHTR/HEI/LAMR/POOL/
#     AON/ALLE/DEO. CVX -35% sold ~$8.4B (84M sh / $17.5B). STZ -95% wiped.
#     GOOGL TRIPLED to $15.6B (54M sh). New: DAL $2.6B, GOOG $1B, Macy's $55M.
#     NYT tripled. LEN +43%. Op earnings $11.35B +18%. Insurance +28%.
#     OXY unchanged at 264.94M sh ($17.22B at $65/sh, +58% YTD).
#     SIRI +5M sh to 124.81M ($2.88B, 37% of SIRI outstanding).
#
# Sector impact matrix (discount rate premium / growth haircut):
#   BENEFICIARIES: Energy ($112 Brent), Defense (war ongoing), Insurance (Hormuz)
#   NEUTRAL:       Healthcare, Utilities, Consumer Staples
#   HURT:          Transportation/Auto ($108 WTI), Industrials (supply chains)
#                  Consumer Disc (oil + tariffs), Tech (valuation rich)

SECTOR_GEO_ADJUSTMENTS = {
    # sector_keyword: (discount_rate_premium, growth_haircut, margin_of_safety_add)
    #
    # v5 (May 22, 2026): Iran war ESCALATING again, Brent $112, WTI $108. Hormuz
    # mostly CLOSED. Stagflation re-emerging. Supreme Court tariff ruling but
    # tariffs persist. Energy benefit STRENGTHENED. Transportation/Auto pain
    # WORSE (fuel cost shock). Insurance premium boosted (Chubb Hormuz lead).
    # S&P still at ATH 7,501 but valuation premium warranted.
    #
    # STRONG Beneficiaries — bigger discount cut, growth boost
    'Energy':                  (-0.025,  0.035, -0.06),   # Brent $112, war ongoing — STRONGEST tailwind
    'Oil & Gas':               (-0.025,  0.035, -0.06),   # CVX/OXY: BRK still ~$35B combined exposure
    'Aerospace & Defense':     (-0.012,  0.030, -0.05),   # Defense budgets accelerating; war ongoing
    'Insurance':               (-0.010,  0.015, -0.02),   # CB Hormuz war premiums STRONG tailwind
    # Neutral / Defensive
    'Healthcare':              (0.003,   0.005, 0.00),    # UNH exited by BRK but defensive overall
    'Utilities':               (0.008,   0.005, 0.00),    # Rate-sensitive; energy cost pass-through
    'Consumer Staples':        (0.012,  -0.008, 0.025),   # Tariff + commodity input cost intensifying
    'Pharmaceuticals':         (0.003,   0.005, 0.00),
    # Mild Hurt — valuation-stretched / rate sensitive
    'Technology':              (0.015,  -0.015, 0.040),   # Fwd P/E 21.2x rich; mixed AI tailwind
    'Semiconductors':          (0.020,  -0.015, 0.040),   # AI demand strong but valuation peaks
    'Software':                (0.010,  -0.010, 0.020),   # AI tailwind partly offsets
    'Financial Services':      (0.015,  -0.012, 0.035),   # Stagflation concern, but credit OK
    'Banks':                   (0.020,  -0.015, 0.040),   # BAC trimmed by BRK; rate uncertainty
    'Real Estate':             (0.022,  -0.020, 0.050),   # Higher-for-longer + oil cost pressure
    # SEVERELY Hurt — oil shock + tariffs
    'Consumer Discretionary':  (0.025,  -0.025, 0.045),   # $112 Brent = $5.00+ gas, tariffs back
    'Retail':                  (0.025,  -0.025, 0.045),   # Macy's added by BRK (cheap entry)
    'Restaurants':             (0.020,  -0.020, 0.035),   # Input cost shock intensifying
    'Automotive':              (0.030,  -0.035, 0.075),   # $108 WTI + tariffs + EV uncertainty
    'Industrial':              (0.020,  -0.020, 0.040),   # Supply chain Hormuz disruption + tariffs
    'Industrials':             (0.020,  -0.020, 0.040),
    'Materials':               (0.012,  -0.012, 0.025),   # Mixed: commodity tailwind / input costs
    'Communication Services':  (0.005,  -0.005, 0.010),   # Insulated; GOOGL tripled by BRK
    'Transportation':          (0.028,  -0.028, 0.060),   # $108 WTI MASSIVE fuel headwind; DAL/airlines
    # ── GICS broad-sector aliases (v6 fallback) ──────────────────────────────
    # yfinance's granular `industry` (e.g. "Medical Devices") rarely matches the
    # keys above, so match_sector_key() falls back to the 11 GICS `sector`
    # values. These aliases ensure every S&P 500 name gets an overlay.
    'Consumer Cyclical':       (0.025,  -0.025, 0.045),   # GICS alias → Consumer Discretionary
    'Consumer Defensive':      (0.012,  -0.008, 0.025),   # GICS alias → Consumer Staples
}

# Base geopolitical risk premium applied to ALL companies
# v5 (May 22): TIGHTENED vs v4 — oil shock ($112 Brent), Iran war escalating,
# stagflation re-emerging, ISM prices at multi-year high. But S&P at fresh ATH.
# Net: discount premium and MOS BUMPED to reflect oil/war reality.
GEO_BASE_DISCOUNT_PREMIUM = 0.013   # +1.3% base discount (v4: 1.0%) — oil shock + stagflation re-emerging
GEO_BASE_GROWTH_HAIRCUT = -0.005    # -0.5% growth haircut (v4: -0.3%) — tariff+oil double pressure
GEO_BASE_MOS_ADD = 0.040            # +4.0% margin of safety (v4: 3.5%) — war ongoing + ATH valuations

# Surfaced in the web dashboard header so users see which overlay generated the numbers.
GEO_VERSION = "v5 (May 22, 2026)"
GEO_NARRATIVE = (
    "Iran war ongoing since Feb 28; Brent $107–112, WTI $102–108; "
    "Hormuz mostly closed. Fed 3.50–3.75%, PCE 2.7% (sticky). "
    "S&P 500 ATH 7,501 (May 15), fwd P/E ~21x. "
    "BRK Q1 2026 13F: $263B portfolio, $397B cash, GOOGL tripled."
)

# DCF defaults — kept here as module-level constants so the runner can export them
# to assumptions.json and the JS recompute uses the identical numbers.
DCF_DISCOUNT_RATE = 0.10
DCF_TERMINAL_GROWTH = 0.025
DCF_PROJECTION_YEARS = 15
DCF_MARGIN_OF_SAFETY = 0.25


# ── Macro snapshot — variable inputs surfaced in the dashboard "Macro lens" ──
# These drive the per-slider guidance shown in the UI. Update alongside the
# geopolitical overlay version (see the v5/v6/... bump checklist in CLAUDE.md).
# Lyn Alden's framework: every IV input is a function of the macro regime;
# the dashboard now shows the macro context inline so the user can decide
# whether the defaults are right for *this* regime.
MACRO_SNAPSHOT = {
    "as_of": "2026-05-22",
    "regime": "Stagflation re-emerging — oil shock + tariff persistence + sticky inflation",
    "brent_usd": "$107–112",
    "wti_usd": "$102–108",
    "fed_funds_pct": "3.50–3.75%",
    "pce_yoy_pct": "2.7%",
    "ism_prices_idx": 84.6,           # highest since Apr 2022
    "sp500_level": 7501,              # ATH May 15
    "sp500_fwd_pe": 21.0,
    "ten_yr_treasury_pct": "~4.5%",
    "recession_prob_wsj_pct": 28,
    "brk_cash_b": 397,                # record
    "geopolitics": "Iran war ongoing since Feb 28; Hormuz mostly closed; fragile ceasefire.",
}

# Per-slider macro guidance (Lyn Alden lens) — what the parameter is, when to
# raise it, when to lower it, and where the current regime points.
SLIDER_GUIDANCE = {
    "discount_premium": {
        "what": "Adds to the 10% base discount rate. Higher → future cash flows are worth less today.",
        "raise_when": "Rising real yields, widening risk premia, geopolitical shocks, late-cycle valuations.",
        "lower_when": "Fed cutting decisively, falling yields, low credit spreads, early-cycle.",
        "today": "Oil shock + sticky 2.7% PCE + ISM prices 84.6 + 10-yr ~4.5% → premium ON. Stagflation argues for a higher real discount rate (default 1.3% is light; 2.5–4% defensible).",
    },
    "growth_haircut": {
        "what": "Subtracts from each ticker's starting growth before the DCF compounds it 15 years.",
        "raise_when": "Stagflation, tariff/supply-chain hits, weakening consumer, energy spikes for non-energy names.",
        "lower_when": "Acceleration, productivity boom, falling input costs, expansionary fiscal/monetary.",
        "today": "Tariff regime persists post-IEEPA ruling; oil hurts margins. Default -0.5% haircut is mild; -1 to -2% reasonable for cyclicals.",
    },
    "mos_add": {
        "what": "Margin of safety added on top of Buffett's base 25%. Bigger MOS = more conservative IV.",
        "raise_when": "Valuations stretched (S&P fwd P/E ≥ 20), late-cycle, regime change in progress, fat tails.",
        "lower_when": "Cheap broad market, panic discounts, high-conviction quality at a known floor.",
        "today": "S&P at ATH 7,501, fwd P/E ~21x, war ongoing. Default +4.0% MOS is appropriate; +6–10% if you want more cushion.",
    },
    "global_growth_shift": {
        "what": "Adds (or subtracts) a constant to every ticker's revenue growth before the DCF.",
        "raise_when": "You believe consensus growth is too pessimistic across the board.",
        "lower_when": "You expect a synchronized slowdown or recession (e.g. shift everything -1 to -3%).",
        "today": "Recession-probability surveys ~28% (WSJ). Try shifting -1% to stress-test, or 0 if you trust per-ticker.",
    },
}

# Plain-English notes for the fixed DCF inputs, rendered in the "Base inputs" panel.
DCF_INPUT_NOTES = {
    "discount_rate": "Cost of equity. 10% is the long-run S&P nominal average. Adjusted up per sector by the geopolitical overlay below.",
    "terminal_growth": "Sustainable long-run growth from year 15 onward. 2.5% sits at the conservative end of nominal GDP — should not exceed long-run nominal GDP.",
    "projection_years": "DCF horizon. Buffett-style 15 years; long enough that growth that fades early is correctly penalized.",
    "margin_of_safety": "Buffett's bargain margin — buy at a discount to fair value. 25% base; the geopolitical overlay adds more in stressed regimes.",
}


def match_sector_key(industry, sector=None):
    """Return the SECTOR_GEO_ADJUSTMENTS key matching this name, or None.

    Tries the granular yfinance `industry` first (e.g. "Oil & Gas Integrated"),
    then falls back to the broad GICS `sector` (e.g. "Healthcare"). The fallback
    is what lifts overlay coverage from ~44% to ~100%: most S&P 500 industry
    strings don't contain an overlay keyword, but every name has a GICS sector.
    """
    for candidate in (industry, sector):
        if not candidate:
            continue
        cl = candidate.lower()
        for sector_key in SECTOR_GEO_ADJUSTMENTS:
            if sector_key.lower() in cl:
                return sector_key
    return None


def get_geo_adjustments(industry, sector=None):
    """Look up sector-specific geopolitical adjustments from industry/sector."""
    sector_key = match_sector_key(industry, sector)
    if sector_key is None:
        return (GEO_BASE_DISCOUNT_PREMIUM, GEO_BASE_GROWTH_HAIRCUT, GEO_BASE_MOS_ADD)
    dr_adj, g_adj, mos_adj = SECTOR_GEO_ADJUSTMENTS[sector_key]
    return (
        GEO_BASE_DISCOUNT_PREMIUM + dr_adj,
        GEO_BASE_GROWTH_HAIRCUT + g_adj,
        GEO_BASE_MOS_ADD + mos_adj,
    )


# ── Buffett-style intrinsic value (v4 — Geopolitical Risk-Adjusted) ──────────
def buffett_intrinsic_value(fcf, growth_rate, liquid_assets=0,
                            discount_rate=0.10,
                            terminal_growth=0.025,
                            projection_years=15,
                            margin_of_safety=0.25,
                            industry=None,
                            sector=None):
    """
    Conservative 15-year Owner Earnings DCF with 2026 geopolitical overlay v5.

    Base model (Buffett philosophy):
    - 15-year FCF projection with 3-phase growth decay
    - Add liquid assets (cash + short-term investments)
    - 25% base margin of safety

    Geopolitical overlay v5 (May 22, 2026) — IRAN WAR + OIL SHOCK:
    - Iran war ongoing since Feb 28 (Khamenei killed). Apr 8 ceasefire fragile.
      Brent $107-112, WTI $102-108 (+45% since war began). Hormuz mostly closed.
      Trump called off planned May 18 strike; Iran reviewing 14-pt framework.
    - Supreme Court (Feb) ruled IEEPA tariffs unconstitutional → ~10% tariffs but
      other statutes preserve broad presidential authority. ISM prices 84.6.
    - Fed 3.50-3.75% (held with 4 dissents). PCE 2.7%. Stagflation re-emerging.
    - S&P 500 ATH 7,501 (May 15). Fwd P/E ~21x. Earnings +15.1% Q1 beat.
    - BRK Q1 2026 13F (5/15): $263B portfolio (-13 positions), cash $397B record.
      16 exits, CVX -35%, STZ -95%, GOOGL TRIPLED, DAL/GOOG/M new positions.
    - Sector-specific: Energy/Defense/Insurance STRONG benefit; Transportation/
      Auto/Industrial heavy hurt from oil shock + tariffs; Tech mild valuation
      haircut; Consumer pain RE-intensified.
    """
    if fcf is None or (isinstance(fcf, float) and pd.isna(fcf)) or fcf <= 0:
        return None
    if growth_rate is None or (isinstance(growth_rate, float) and pd.isna(growth_rate)):
        growth_rate = 0.03
    if liquid_assets is None or (isinstance(liquid_assets, float) and pd.isna(liquid_assets)):
        liquid_assets = 0
    liquid_assets = max(liquid_assets, 0)

    # Apply geopolitical adjustments
    geo_dr, geo_g, geo_mos = get_geo_adjustments(industry, sector)
    adj_discount = discount_rate + geo_dr
    adj_mos = margin_of_safety + geo_mos
    growth_rate = growth_rate + geo_g

    # Cap starting growth conservatively (wider band for crisis)
    g_start = min(max(growth_rate, -0.05), 0.12)
    g_terminal = terminal_growth

    dcf_sum = 0.0
    projected_fcf = float(fcf)

    for yr in range(1, projection_years + 1):
        if yr <= 5:
            g = g_start
        elif yr <= 10:
            fade = (yr - 5) / 5.0
            g = g_start * (1 - fade) + g_terminal * fade
        else:
            g = g_terminal

        projected_fcf *= (1 + g)
        dcf_sum += projected_fcf / ((1 + adj_discount) ** yr)

    terminal_fcf = projected_fcf * (1 + g_terminal)
    terminal_value = terminal_fcf / (adj_discount - g_terminal)
    terminal_pv = terminal_value / ((1 + adj_discount) ** projection_years)

    enterprise_value = dcf_sum + terminal_pv + liquid_assets

    return enterprise_value * (1 - adj_mos)


# ── Financials: FCF-DCF is meaningless for banks/insurers (float, leverage) ──
# These keywords route a name to the justified-P/B excess-return model instead.
FINANCIAL_KEYWORDS = (
    'bank', 'insurance', 'capital markets', 'asset management', 'financial',
    'credit services', 'stock exchanges', 'brokerage', 'mortgage',
    'reit', 'real estate',
)


def is_financial(industry, sector=None):
    """True if the name should use the book-value model rather than FCF-DCF."""
    for candidate in (industry, sector):
        if candidate and any(k in candidate.lower() for k in FINANCIAL_KEYWORDS):
            return True
    return False


def financial_intrinsic_value(book_value_per_share, shares_out, roe,
                              discount_rate=0.10,
                              terminal_growth=0.025,
                              margin_of_safety=0.25,
                              industry=None,
                              sector=None):
    """Justified-P/B excess-return fair value for financials & REITs.

    Fair P/B = (ROE − g) / (r − g), then fair equity = fair P/B × book equity.
    This replaces FCF-DCF, which both (a) returns nothing for banks (no
    meaningful FCF) and (b) wildly overvalues insurers (float inflates FCF —
    e.g. PRU showed +440% under DCF). ROE and the P/B multiple are bounded to
    keep outputs sane. Reuses the same geopolitical overlay for r and MOS so
    the what-if sliders move financials too.
    """
    if book_value_per_share is None or shares_out is None or roe is None:
        return None
    if isinstance(book_value_per_share, float) and pd.isna(book_value_per_share):
        return None
    if isinstance(roe, float) and pd.isna(roe):
        return None
    book_equity = book_value_per_share * shares_out
    if book_equity <= 0:
        return None

    geo_dr, geo_g, geo_mos = get_geo_adjustments(industry, sector)
    r = discount_rate + geo_dr
    g = max(terminal_growth + geo_g, 0.0)          # sliders nudge sustainable growth
    if r - g < 0.02:                                # floor the spread to avoid blowups
        r = g + 0.02
    roe = min(max(roe, -0.05), 0.30)                # clamp distorted reported ROE
    adj_mos = margin_of_safety + geo_mos

    fair_pb = (roe - g) / (r - g)
    fair_pb = min(max(fair_pb, 0.2), 4.0)           # sane P/B band
    fair_equity = fair_pb * book_equity
    return fair_equity * (1 - adj_mos)


# ── Fetch with retries ────────────────────────────────────────────────────────
def fetch_single_ticker(ticker):
    """Fetch data for one ticker with retries. Returns (data_dict, error_msg)."""
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            t = yf.Ticker(ticker)
            info = t.info or {}

            if not info or info.get('regularMarketPrice') is None and info.get('currentPrice') is None:
                # Sometimes yfinance returns empty info on first try
                if attempt < MAX_RETRIES:
                    time.sleep(BASE_DELAY * attempt)
                    continue
                return {k: None for k in ['price','yr_change','mcap_b','rev_growth','gross_margin','pe','fcf_m','iv_b','discount','p_fcf']}, f"Empty info after {MAX_RETRIES} attempts"

            price = info.get('currentPrice') or info.get('regularMarketPrice')
            price = validate('price', price)

            yr_change = info.get('52WeekChange')
            if yr_change is None:
                try:
                    hist = t.history(period='1y')
                    if not hist.empty and price:
                        first_close = hist['Close'].dropna().iloc[0]
                        if first_close > 0:
                            yr_change = (price - first_close) / first_close
                except Exception:
                    pass

            mcap = info.get('marketCap')
            mcap_b = validate('mcap_b', mcap / 1e9 if mcap else None)

            rev_growth = validate('rev_growth', info.get('revenueGrowth'))
            gross_margin = validate('gross_margin', info.get('grossMargins'))
            pe = validate('pe', info.get('trailingPE') or info.get('forwardPE'))

            fcf = info.get('freeCashflow')
            fcf_m = validate('fcf_m', fcf / 1e6 if fcf else None)

            # P/FCF ratio
            p_fcf = None
            if price and fcf and info.get('sharesOutstanding'):
                fcf_per_share = fcf / info['sharesOutstanding']
                if fcf_per_share != 0:
                    p_fcf = validate('p_fcf', price / fcf_per_share)

            # Liquid assets: cash + short-term investments on balance sheet
            cash = info.get('totalCash') or 0  # cash & cash equivalents
            # yfinance 'totalCash' includes cash + short-term investments

            # Industry (granular) + GICS sector (broad fallback) for overlay
            industry = info.get('industry') or ''
            sector = info.get('sector') or ''
            shares_out = info.get('sharesOutstanding')

            # Book-value inputs (financials valuation)
            bvps = info.get('bookValue')          # book value PER SHARE
            roe = info.get('returnOnEquity')

            # Route financials/REITs to the P/B model; everyone else to FCF-DCF.
            if is_financial(industry, sector):
                valuation_method = 'pb'
                iv = financial_intrinsic_value(
                    bvps, shares_out, roe,
                    industry=industry, sector=sector,
                )
            else:
                valuation_method = 'dcf'
                iv = buffett_intrinsic_value(
                    fcf,
                    rev_growth if rev_growth else 0.03,
                    liquid_assets=cash,
                    industry=industry,
                    sector=sector,
                )
            iv_b = iv / 1e9 if iv else None

            discount = None
            if iv_b and mcap_b and mcap_b > 0:
                discount = (iv_b - mcap_b) / mcap_b

            return {
                'price': price,
                'yr_change': yr_change,
                'mcap_b': mcap_b,
                'rev_growth': rev_growth,
                'gross_margin': gross_margin,
                'pe': pe,
                'fcf_m': fcf_m,
                'iv_b': iv_b,
                'discount': discount,
                'p_fcf': p_fcf,
                'fcf': fcf,                                  # raw $ — needed by JS recompute
                'cash': cash,                                # raw $ — needed by JS recompute
                'shares_out': shares_out,                    # raw — needed by JS recompute
                'industry': industry,                        # for sector matching in JS
                'sector': sector,                            # GICS broad fallback
                'bvps': bvps,                                # financials recompute
                'roe': roe,                                  # financials recompute
                'valuation_method': valuation_method,        # 'dcf' | 'pb'
            }, None

        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"
            if attempt < MAX_RETRIES:
                time.sleep(BASE_DELAY * (2 ** (attempt - 1)))

    return {k: None for k in [
        'price','yr_change','mcap_b','rev_growth','gross_margin','pe','fcf_m',
        'iv_b','discount','p_fcf','fcf','cash','shares_out','industry',
        'sector','bvps','roe','valuation_method'
    ]}, last_err


def fetch_batch(tickers):
    """Fetch data for a list of tickers, isolating per-ticker errors."""
    results = {}
    errors = {}
    for ticker in tickers:
        data, err = fetch_single_ticker(ticker)
        results[ticker] = data
        if err:
            errors[ticker] = err
    return results, errors


# ── Per-company intrinsic-value models (portfolio) ───────────────────────────
# Loaded from portfolio_models.json. Two conservative-Buffett model types:
#   fcf_dcf            — 15-yr mid-year per-share owner-earnings DCF (reproduces
#                        the user's Google-Sheet models within <0.5%).
#   earnings_multiple  — normalized EPS x justified P/E (+ net cash per share),
#                        for banks / high-leverage names where an FCF DCF with
#                        full net-debt subtraction is the wrong tool.
PORTFOLIO_MODELS_PATH = os.path.join(SCRIPT_DIR, 'portfolio_models.json')


def model_iv_per_share(spec):
    """Compute intrinsic value per share (in the listing/price currency) from a
    model spec. Returns None for non-modelable specs (e.g. ETF)."""
    mtype = spec.get("model_type")
    if mtype == "fcf_dcf":
        g = spec["growth_ps"]
        fcf = spec["fcf_ps0"]
        r = spec["discount_rate"]
        tg = spec["terminal_growth"]
        pv = 0.0
        for n in range(1, 16):
            gn = g[0] if n <= 5 else g[1] if n <= 10 else g[2]
            fcf *= (1 + gn)
            pv += fcf / (1 + r) ** (n - 0.5)
        tv = fcf * (1 + tg) / (r - tg)
        pv_tv = tv / (1 + r) ** (15 - 0.5)
        return pv + pv_tv - spec.get("net_debt_ps", 0.0)
    if mtype == "earnings_multiple":
        return spec["normalized_eps"] * spec["target_pe"] + spec.get("net_cash_ps", 0.0)
    if mtype == "netcash":
        # IV is a precomputed net-cash floor (data too thin to recompute live).
        return spec.get("iv_per_share")
    return None


def load_portfolio_models():
    if not os.path.exists(PORTFOLIO_MODELS_PATH):
        return {}
    try:
        with open(PORTFOLIO_MODELS_PATH) as f:
            return json.load(f).get("models", {})
    except Exception as e:
        log(f"portfolio models: could not read {PORTFOLIO_MODELS_PATH}: {e}")
        return {}


def fetch_portfolio_rows(holdings, models=None):
    """Fetch + value an explicit watchlist that may include non-S&P, foreign,
    or ETF symbols. Returns rows in the SAME shape as LAST_RUN_ROWS, each
    flagged in_portfolio=True.

    Valuation: if a per-company model exists in portfolio_models.json, the
    intrinsic value comes from THAT model (recomputed from its inputs) and the
    discount is (IV/share - live price) / live price — a clean same-currency
    ratio. Otherwise it falls back to the generic DCF/P-B engine. ETFs / names
    with no model + no usable FCF get iv_b=None (price only)."""
    if models is None:
        models = load_portfolio_models()
    rows = []
    errors = {}
    for h in holdings:
        sym = h["symbol"]
        data, err = fetch_single_ticker(sym)
        if err:
            errors[sym] = err
        industry_str = data.get("industry") or ""
        sector_str = data.get("sector") or ""
        row = {
            "ticker": sym,
            "company": h.get("display") or sym,
            "industry": industry_str,
            "sector": sector_str,
            "sector_key": match_sector_key(industry_str, sector_str),
            "valuation_method": data.get("valuation_method"),
            "price": data.get("price"),
            "yr_change": data.get("yr_change"),
            "mcap_b": data.get("mcap_b"),
            "rev_growth": data.get("rev_growth"),
            "gross_margin": data.get("gross_margin"),
            "pe": data.get("pe"),
            "fcf_m": data.get("fcf_m"),
            "iv_b": data.get("iv_b"),
            "discount": data.get("discount"),
            "p_fcf": data.get("p_fcf"),
            "fcf": data.get("fcf"),
            "cash": data.get("cash"),
            "shares_out": data.get("shares_out"),
            "bvps": data.get("bvps"),
            "roe": data.get("roe"),
            "brk_held": "—",
            "brk_pos_b": None,
            "in_portfolio": True,
            "sp500": False,   # non-constituent (foreign/ADR/ETF) — exclude from S&P scans
        }
        apply_model_to_row(row, models.get(sym))
        rows.append(row)
    return rows, errors


def apply_model_to_row(row, spec):
    """If a portfolio model spec exists, override the row's IV/discount with the
    model-derived intrinsic value per share vs the live price. Adds model_*
    fields for the dashboard/digest. No-op if spec is None or non-modelable."""
    if not spec:
        return
    row["model_source"] = spec.get("source")
    row["model_confidence"] = spec.get("confidence")
    row["model_note"] = spec.get("note")
    row["model_currency"] = spec.get("currency")
    row["model_discount_rate"] = spec.get("discount_rate") or spec.get("discount_rate_ref")
    row["model_terminal_growth"] = spec.get("terminal_growth")
    row["sheet_iv_per_share"] = spec.get("sheet_iv_per_share")
    inputs = spec.get("inputs") or {}
    row["model_growth_y1_5"] = inputs.get("fcf_growth_y1_5")
    row["model_analyst_guard"] = inputs.get("analyst_guard")
    if spec.get("model_type") == "none":
        row["valuation_method"] = "etf"
        row["iv_per_share"] = None
        return
    iv_ps = model_iv_per_share(spec)
    if iv_ps is None:
        return
    row["valuation_method"] = "model"
    row["model_type"] = spec.get("model_type")
    row["iv_per_share"] = round(iv_ps, 4)
    price = row.get("price")
    if price and price > 0:
        row["discount"] = (iv_ps - price) / price
        # iv_b in the listing currency, consistent with yfinance mcap_b for
        # the same ticker (both local for foreign, both USD for US/ADR).
        if row.get("mcap_b") is not None and price:
            shares_implied = row["mcap_b"] / price  # billions of shares
            row["iv_b"] = iv_ps * shares_implied
    else:
        row["discount"] = None


# ── Style helpers ─────────────────────────────────────────────────────────────
GREEN_FONT = Font(name='Arial', size=9, bold=True, color='006100')
RED_FONT = Font(name='Arial', size=9, bold=True, color='9C0006')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
NUM_FONT = Font(name='Arial', size=9)
GRAY_FONT = Font(name='Arial', size=9, color='CCCCCC')
CENTER = Alignment(horizontal='center')
ALT1 = PatternFill('solid', fgColor='F2F7FB')
ALT2 = PatternFill('solid', fgColor='FFFFFF')
BORDER = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)

def write_cell(ws, row, col, val, fmt, conditional, base_fill):
    cell = ws.cell(row=row, column=col)
    if val is not None and not (isinstance(val, float) and pd.isna(val)):
        cell.value = val
        cell.number_format = fmt
        if conditional and isinstance(val, (int, float)):
            cell.font = GREEN_FONT if val > 0 else RED_FONT if val < 0 else NUM_FONT
            cell.fill = GREEN_FILL if val > 0 else RED_FILL if val < 0 else base_fill
        else:
            cell.font = NUM_FONT
            cell.fill = base_fill
        return True
    else:
        cell.value = '—'
        cell.font = GRAY_FONT
        cell.fill = base_fill
        return False
    cell.alignment = CENTER
    cell.border = BORDER


# ── Run-row collection (exposed to the dashboard runner) ─────────────────────
# Populated during each update_dashboard() run with one entry per surviving
# ticker. The web update agent reads this instead of re-opening Portfolio.xlsx.
LAST_RUN_ROWS = []


def get_last_run_rows():
    return list(LAST_RUN_ROWS)


def clear_last_run_rows():
    LAST_RUN_ROWS.clear()


# ── Main update ───────────────────────────────────────────────────────────────
def update_dashboard():
    start_time = time.time()
    log("=" * 60)
    log("S&P 500 Dashboard update starting...")
    clear_last_run_rows()

    if not os.path.exists(XLSX_PATH):
        log(f"ERROR: {XLSX_PATH} not found")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)
    if 'Dashboard' not in wb.sheetnames:
        log("ERROR: 'Dashboard' sheet not found")
        sys.exit(1)

    ws = wb['Dashboard']

    # Read tickers
    tickers = []
    row = DATA_START
    while True:
        val = ws.cell(row=row, column=2).value
        if val is None or str(val).strip() == '':
            break
        tickers.append((row, str(val).strip()))
        row += 1

    log(f"Found {len(tickers)} tickers")

    # Fetch in batches
    all_data = {}
    all_errors = {}
    ticker_list = [t for _, t in tickers]

    for i in range(0, len(ticker_list), BATCH_SIZE):
        batch = ticker_list[i:i + BATCH_SIZE]
        batch_n = i // BATCH_SIZE + 1
        total_b = (len(ticker_list) + BATCH_SIZE - 1) // BATCH_SIZE
        log(f"  Batch {batch_n}/{total_b} ({len(batch)} tickers)...")

        data, errs = fetch_batch(batch)
        all_data.update(data)
        all_errors.update(errs)

        if i + BATCH_SIZE < len(ticker_list):
            time.sleep(2)

    # Write MARKET data only — columns E(5) through N(14)
    # BRK columns O-S are PRESERVED (static 13F data, not overwritten)
    # E=price F=yr_change G=mcap H=rev_growth I=gross_margin J=pe K=fcf L=iv M=discount N=p_fcf
    success = 0
    partial = 0
    failed = 0
    no_data_tickers = []  # tickers with zero data (candidates for removal)

    for row_num, ticker in tickers:
        d = all_data.get(ticker, {})
        base_fill = ALT1 if (row_num - DATA_START) % 2 == 0 else ALT2

        col_map = [
            (5,  d.get('price'),        '$#,##0.00', False),
            (6,  d.get('yr_change'),     '0.0%',     True),
            (7,  d.get('mcap_b'),        '#,##0.0',  False),
            (8,  d.get('rev_growth'),    '0.0%',     True),
            (9,  d.get('gross_margin'),  '0.0%',     False),
            (10, d.get('pe'),            '0.0',      False),
            (11, d.get('fcf_m'),         '#,##0',    False),
            (12, d.get('iv_b'),          '#,##0.0',  False),
            (13, d.get('discount'),      '0.0%',     True),
            (14, d.get('p_fcf'),         '0.0',      False),
        ]

        filled = 0
        for col, val, fmt, cond in col_map:
            ok = write_cell(ws, row_num, col, val, fmt, cond, base_fill)
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = CENTER
            cell.border = BORDER
            if ok:
                filled += 1

        if filled >= 7:
            success += 1
        elif filled > 0:
            partial += 1
        else:
            failed += 1
            no_data_tickers.append((row_num, ticker))

        # Collect a row for the web update agent. Skip tickers that produced
        # no data at all — they'll be deleted from the sheet below.
        if filled > 0:
            company = ws.cell(row=row_num, column=3).value
            industry_col = ws.cell(row=row_num, column=4).value
            brk_held = ws.cell(row=row_num, column=15).value   # column O
            brk_pos_b = ws.cell(row=row_num, column=16).value  # column P
            industry_str = d.get('industry') or industry_col or ''
            sector_str = d.get('sector') or ''
            LAST_RUN_ROWS.append({
                'ticker': ticker,
                'company': company,
                'industry': industry_str,
                'sector': sector_str,
                'sector_key': match_sector_key(industry_str, sector_str),
                'valuation_method': d.get('valuation_method'),
                'price': d.get('price'),
                'yr_change': d.get('yr_change'),
                'mcap_b': d.get('mcap_b'),
                'rev_growth': d.get('rev_growth'),
                'gross_margin': d.get('gross_margin'),
                'pe': d.get('pe'),
                'fcf_m': d.get('fcf_m'),
                'iv_b': d.get('iv_b'),
                'discount': d.get('discount'),
                'p_fcf': d.get('p_fcf'),
                'fcf': d.get('fcf'),
                'cash': d.get('cash'),
                'shares_out': d.get('shares_out'),
                'bvps': d.get('bvps'),
                'roe': d.get('roe'),
                'brk_held': brk_held if brk_held not in (None, '') else '—',
                'brk_pos_b': brk_pos_b if isinstance(brk_pos_b, (int, float)) else None,
            })

    # ── No-data tickers: NON-DESTRUCTIVE (v6) ─────────────────────────────────
    # Previously these rows were deleted from the sheet, which on a flaky-network
    # run would permanently destroy the manual BRK columns O-S for that ticker.
    # Now we keep the row: A-D (ticker/company/industry) and O-S (BRK) are
    # untouched; E-N already show '—' from write_cell. A transient fetch failure
    # is recoverable on the next run instead of silently shrinking the index.
    no_data_count = len(no_data_tickers)
    if no_data_tickers:
        log(f"{no_data_count} companies returned no market data this run "
            f"(rows + BRK columns PRESERVED, not deleted):")
        for row_num, ticker in no_data_tickers:
            log(f"  {ticker} (row {row_num}) — no data this run, kept")

    # Update timestamp in A2
    now = datetime.datetime.now()
    elapsed = time.time() - start_time
    ws['A2'] = (
        f'Last updated: {now:%Y-%m-%d %H:%M:%S} | '
        f'{success} full / {partial} partial / {no_data_count} no-data (kept) | '
        f'{elapsed:.0f}s elapsed'
    )
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    # ── Write Errors sheet ────────────────────────────────────────────────────
    if 'Errors' in wb.sheetnames:
        del wb['Errors']
    if all_errors:
        es = wb.create_sheet('Errors')
        es['A1'] = 'Ticker'
        es['B1'] = 'Error'
        es['C1'] = 'Timestamp'
        for c in range(1, 4):
            es.cell(row=1, column=c).font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            es.cell(row=1, column=c).fill = PatternFill('solid', fgColor='C00000')
        for idx, (tk, err) in enumerate(all_errors.items(), 2):
            es.cell(row=idx, column=1, value=tk).font = Font(name='Arial', size=9)
            es.cell(row=idx, column=2, value=err).font = Font(name='Arial', size=9)
            es.cell(row=idx, column=3, value=now.strftime('%Y-%m-%d %H:%M:%S')).font = Font(name='Arial', size=9)
        es.column_dimensions['A'].width = 10
        es.column_dimensions['B'].width = 60
        es.column_dimensions['C'].width = 20

    # ── Update auto-filter to cover all data rows ────────────────────────────
    last_row = DATA_START
    while ws.cell(row=last_row, column=2).value:
        last_row += 1
    last_row -= 1
    if last_row >= DATA_START:
        ws.auto_filter.ref = f'A{HEADER_ROW}:S{last_row}'

    wb.save(XLSX_PATH)

    # ── Coverage stats for run-health logging ─────────────────────────────────
    iv_count = sum(1 for r in LAST_RUN_ROWS if r.get('iv_b') is not None)
    sector_count = sum(1 for r in LAST_RUN_ROWS if r.get('sector_key'))
    n_rows = len(LAST_RUN_ROWS) or 1

    # ── JSON report ───────────────────────────────────────────────────────────
    report = {
        'timestamp': now.isoformat(),
        'elapsed_seconds': round(elapsed, 1),
        'total_tickers': len(tickers),
        'success': success,
        'partial': partial,
        'no_data_kept': no_data_count,
        'errors_count': len(all_errors),
        'iv_coverage': round(iv_count / n_rows, 4),
        'sector_coverage': round(sector_count / n_rows, 4),
        'no_data_tickers': [t for _, t in no_data_tickers],
        'sample_errors': dict(list(all_errors.items())[:10]),
    }
    try:
        with open(REPORT_PATH, 'w') as f:
            json.dump(report, f, indent=2)
    except Exception:
        pass

    log(f"Update complete: {success} full / {partial} partial / {no_data_count} no-data (kept) | {len(all_errors)} errors")
    log(f"Coverage: IV {100*iv_count//n_rows}% | sector overlay {100*sector_count//n_rows}%")
    log(f"Elapsed: {elapsed:.1f}s | Saved to {XLSX_PATH}")
    log("=" * 60)

    return report


if __name__ == '__main__':
    report = update_dashboard()
    print(json.dumps(report, indent=2))
